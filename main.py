import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import threading
import time
import os
import psutil
import re
import openpyxl
import customtkinter as ctk
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException
from datetime import datetime

"""
Automação de mensagem padrão de prorrogação de contrato de experiencia no Onvio Messenger.
"""

# Configuração do tema do customtkinter
ctk.set_appearance_mode("System")  # Modos: "System" (padrão), "Dark", "Light"
ctk.set_default_color_theme("blue")  # Temas: "blue" (padrão), "green", "dark-blue"

# Variável global para controlar o cancelamento do processamento
cancelar = False
log_file_path = None

#FUNÇÕES DO PROGRAMA
def focar_barra_endereco_e_navegar(driver, termo_busca):
    try:
        
        time.sleep(1)
        # Localiza e clica barra de contato
        focused_element = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="page-content"]/div/div[2]/div[3]/div[1]/div/div/div/input'))
        )
        
            
        # Simula a pressão da tecla Tab várias vezes
        # for i in range(15):  # Ajuste este número conforme necessário
            
        # Verifica se o elemento focado é a barra de pesquisa
        # focused_element = driver.switch_to.active_element
        if focused_element.get_attribute('placeholder') == "Buscar contatos...":
            focused_element.click()
            atualizar_log(f"Verificando contato {termo_busca}...")
            
            
            # Tenta preencher o campo
            try:
                # Verifica se o texto foi inserido corretamente
                valor_atual = focused_element.get_attribute('value')
                if termo_busca != valor_atual:
                    focused_element.clear()  # Limpa o campo primeiro
                    focused_element.send_keys(termo_busca)
                    atualizar_log(f"Texto '{termo_busca}' inserido com sucesso na barra de pesquisa.")
                    time.sleep(1)  # Espera um pouco para o texto ser inserido
                
                
                elif valor_atual == termo_busca:
                    atualizar_log(f"Texto '{termo_busca}' ja presente na barra de pesquisa.")
                else:
                    atualizar_log(f"Texto inserido não corresponde. Valor atual: '{valor_atual}'")
            except Exception as e:
                atualizar_log(f"Erro ao preencher o campo: {str(e)}")
            
            
            return True
        elif focused_element.get_attribute('placeholder') == "Buscar grupos...":
            atualizar_log(f"Barra de pesquisa encontrada")
            
            
            # Tenta preencher o campo
            try:
                valor_atual = focused_element.get_attribute('value')
                if termo_busca != valor_atual:
                    focused_element.clear()  # Limpa o campo primeiro
                    focused_element.send_keys(termo_busca)
                    time.sleep(1)  # Espera um pouco para o texto ser inserido
                
                # Verifica se o texto foi inserido corretamente
                
                elif valor_atual == termo_busca:
                    atualizar_log(f"Texto '{termo_busca}' inserido com sucesso na barra de pesquisa.")
                else:
                    atualizar_log(f"Texto inserido não corresponde. Valor atual: '{valor_atual}'")
            except Exception as e:
                atualizar_log(f"Erro ao preencher o campo: {str(e)}")
            
            
            return True
                
            
            
            # ActionChains(driver).send_keys(Keys.TAB).perform()
            # time.sleep(1)
            # atualizar_log(f"TAB {i+1} pressionado")
            
        atualizar_log("Barra de pesquisa não encontrada após 10 TABs")
        return False


    except Exception as e:
        atualizar_log(f"Erro ao focar na barra de endereço ou navegar: {str(e)}")
        return False

def encerrar_processos_chrome():
    for proc in psutil.process_iter(['name']):
        if proc.info['name'] == 'chrome.exe':
            try:
                proc.terminate()
            except (psutil.NoSuchProcess, psutil.AccessDenied):
                pass
    time.sleep(2)  # Aguarda um pouco para garantir que os processos foram encerrados

def abrir_chrome_com_url(url):
    encerrar_processos_chrome()
    
    # Caminho para o perfil padrão do Chrome
    user_data_dir = os.path.expanduser('~') + r'\AppData\Local\Google\Chrome\User Data'
    
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument(f"user-data-dir={user_data_dir}")
    chrome_options.add_argument("--profile-directory=Default")
    chrome_options.add_argument("--disable-translate")  # Tenta desabilitar a tradução automática
    chrome_options.add_argument("--lang=pt-BR")  # Define o idioma para português do Brasil
    chrome_options.add_argument("--enable-javascript")
    chrome_options.add_experimental_option("prefs", {
        "translate": {"enabled": "false"},
        "profile.default_content_setting_values.notifications": 2
    })
    
    
    service = Service(ChromeDriverManager().install())
    
    try:
        driver = webdriver.Chrome(service=service, options=chrome_options)
        driver.set_page_load_timeout(180)  # Aumenta o tempo limite para carregar a página
        driver.get(url)
        atualizar_log(f"Chrome aberto com a URL: {url}")
        return driver
    except Exception as e:
        atualizar_log(f"Erro ao abrir o Chrome: {str(e)}")
        return None

def esperar_carregamento_completo(driver):
    global cancelar
    cancelar = False
    try:
        WebDriverWait(driver, 60).until(
            lambda d: d.execute_script('return document.readyState') == 'complete'
        )
        atualizar_log("Página completamente carregada.")
    except TimeoutException as e:
        atualizar_log(f"Erro ao esperar o carregamento completo: {str(e)}")

def processar_resultados_busca(driver):
    global cancelar
    cancelar = False
    try:
        # Espera pelos resultados da busca
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="page-content"]/div/div[2]/div[3]/div[2]/div/div[1]'))
        )
        
        
        # Localiza e tenta clicar no elemento especificado(1º CONTATO)
        elemento_alvo = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="page-content"]/div/div[2]/div[3]/div[2]/div/div[1]'))
        )
        
        # Tenta clicar no elemento
        if elemento_alvo:
            elemento_alvo.click()
            atualizar_log("Clicado no elemento alvo.")
        else:
            atualizar_log("Elemento não encontrado.")
        
        
    except TimeoutException:
        atualizar_log("Timeout ao esperar pelo elemento alvo.", cor='vermelho')
    except Exception as e:
        atualizar_log(f"Erro ao processar resultados da busca ou clicar no elemento: {str(e)}")

def focar_barra_mensagem_enviar(driver, mensagem):
    global cancelar
    if cancelar:
        atualizar_log("Processamento cancelado!", cor="azul")
        return
    try:
        elemento_alvo = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="preview-root"]/div[2]/div[3]/div[1]/div/div[2]/div[2]/div[1]'))
        )
        
        if elemento_alvo.get_attribute('data-placeholder') == "Mensagem":
            elemento_alvo.click()
            # atualizar_log(f"Barra de mensagem encontrada após {i+1} TABs")
            atualizar_log(f"Barra de Mensagem encontrada e clicada!")
            atualizar_log(f"Enviando Mensagem...")
            
            # Tenta preencher o campo
            try:
                    # Divide a mensagem em parágrafos
                paragrafos = re.split(r'\n+', mensagem.strip())

                # Insere cada parágrafo como um novo elemento <p>
                for i, paragrafo in enumerate(paragrafos):
                    if i > 0:
                        # Para todos exceto o primeiro parágrafo, pressiona Shift+Enter para criar um novo <p>
                        ActionChains(driver).key_down(Keys.SHIFT).send_keys(Keys.ENTER).key_up(Keys.SHIFT).perform()
                        time.sleep(0.5)
                        
                    if cancelar:
                        atualizar_log("Processamento cancelado!", cor="azul")
                        return
                    # Insere o texto do parágrafo
                    ActionChains(driver).send_keys(paragrafo).perform()
                    time.sleep(0.5)

                atualizar_log("Mensagem formatada inserida com sucesso.")
                if cancelar:
                    atualizar_log("Processamento cancelado!", cor="azul")
                    return
                # Agora vamos clicar no botão de enviar
                # Primeiro, tentamos localizar o botão por XPath
                try:
                    botao_enviar = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.XPATH, '//*[@id="preview-root"]/div[2]/div[3]/div[3]/div[1]/button'))
                    )
                    botao_enviar.click()
                    atualizar_log("Botão de enviar clicado com sucesso.")
                except:
                    atualizar_log("\nNão foi possível clicar no botão de enviar por XPath. Tentando por Tab.\n", cor="vermelho")
                    
                    return False

                atualizar_log("Desconsiderando mensagem...")
                if cancelar:
                    atualizar_log("Processamento cancelado!", cor="azul")
                    return
                # Agora vamos clicar no botão de desconsiderar
                # Primeiro, tentamos localizar o botão por XPath
                try:
                    botao_desconsiderar = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.XPATH, '//*[@id="ChatHeader"]/div[2]/div[1]/div[3]/div[1]/button/div'))
                    )
                    botao_desconsiderar.click()
                    atualizar_log("Botão de DESCONSIDERAR clicado com sucesso.")
                    time.sleep(2)
                except:
                    atualizar_log("\nNão foi possível clicar no botão de DESCONSIDERAR por XPath. Tentando por Tab.\n", cor="vermelho")
                    
                    return False                

                try:
                    # Espera pelo modal de confirmação
                    WebDriverWait(driver, 5).until(
                        EC.presence_of_element_located((By.XPATH, '/html/body/div[4]'))
                )
                    desconsiderar = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.XPATH, '/html/body/div[4]/div/div/div[3]/button[2]'))
                    )
                    desconsiderar.click()
                    time.sleep(4)
                    atualizar_log("Mensagem Desconsiderada com Sucesso!", cor="azul")
                except:
                    atualizar_log("\nNão foi possível DESCONSIDERAR por XPath. Tentando por Tab.\n", cor="vermelho")
                    
                    return False
                

                
                time.sleep(2)  # Espera um pouco para a mensagem ser enviada
                
                return True
                
            except Exception as e:
                atualizar_log(f"Erro ao preencher o campo ou enviar a mensagem: {str(e)}", cor="vermelho")
            
            # break  # Sai do loop se encontrou e interagiu com a barra de mensagem
                    
        else:  # Se o loop terminar sem encontrar a barra de mensagem
            atualizar_log(f"Barra de mensagem não encontrada após {i+1} TABs.")
        
    except Exception as e:
        atualizar_log(f"Erro ao focar na barra de mensagem ou enviar: {str(e)}")
    
    return False

def encontrar_e_clicar_barra_contatos(driver, contato, grupo):
    global cancelar
    if cancelar:
        atualizar_log("Processamento cancelado!", cor="azul")
        return
    try:
        esperar_carregamento_completo(driver)
        atualizar_log("Aguardando elementos específicos da página...")

        time.sleep(5)
        atualizar_log("Focando na aba de contato correspondente...")  
        
    # ABA GRUPO         
        if grupo.upper() != "NONE":
            if cancelar:
                atualizar_log("Processamento cancelado!", cor="azul")
                return
            focar_pagina(driver)
            
            if focar_barra_endereco_e_navegar(driver, grupo):
                if cancelar:
                    atualizar_log("Processamento cancelado!", cor="azul")
                    return
                atualizar_log("Navegação aba grupo.")
                processar_resultados_busca(driver)
                atualizar_log("Navegação bem-sucedida. Busca, ABA GRUPO, realizada.")
                return True
             
    # ABA CONTATO  
        elif contato.upper() != "NONE".upper():
            if cancelar:
                atualizar_log("Processamento cancelado!", cor="azul")
                return
            
            atualizar_log("Navegação aba Contato.")
            time.sleep(1)
            clicar_voltar_lista_contatos(driver) #//*[@id="react-tabs-0"]
            if focar_barra_endereco_e_navegar(driver, contato):
                if cancelar:
                    atualizar_log("Processamento cancelado!", cor="azul")
                    return
                processar_resultados_busca(driver)
                atualizar_log("Navegação bem-sucedida. Busca, ABA CONTATO, realizada.") 
                return True
        else:
            atualizar_log("Falha na navegação ou busca.", cor="vermelho")
            return False

    except Exception as e:
        atualizar_log(f"Erro geral ao tentar interagir com a página: {str(e)}", cor="vermelho")
        return False
   
def clicar_voltar_lista_contatos(driver):
    global cancelar
    if cancelar:
        atualizar_log("Processamento cancelado!", cor="azul")
        return
    try:
        #xpath perto da barra de contados
        elemento = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="react-tabs-0"]'))
        )
        
        if elemento:
            elemento.click()
        
        atualizar_log("Clicado na aba Contatos.\n", cor="azul")
        time.sleep(1)  # Espera para a página carregar
        return True
    except Exception as e:
        atualizar_log(f"Erro ao clicar no botão para voltar à lista de contatos: {str(e)}", cor="vermelho")
        return False

def focar_pagina(driver):
    global cancelar
    if cancelar:
        atualizar_log("Processamento cancelado!", cor="azul")
        return
    try:    
        # focar na aba grupos dos contatos
        elemento = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="react-tabs-2"]')) 
        )

        if elemento:
            elemento.click()
        

        atualizar_log("Clicado na aba Grupo.", cor="azul")
        time.sleep(3)  # Espera para a página carregar
        return True
    except Exception as e:
        atualizar_log(f"Erro ao clicar no botão para voltar à lista de contatos: {str(e)}", cor="vermelho")
        return False

def focar_pagina_geral(driver):
    global cancelar
    if cancelar:
        atualizar_log("Processamento cancelado!", cor="azul")
        return
    try:    
        # focar na aba contatos Geral
        elemento = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="page-content"]/div/div[2]/div[1]/div/div/div/div[1]/div/div[1]'))
        )
        
        elemento.click()
        atualizar_log("Clicado no contato geral.")
        
        time.sleep(5)  # Espera para a página carregar
        return True
    except Exception as e:
        atualizar_log(f"Erro ao clicar no botão para voltar à lista de contatos: {str(e)}", cor="vermelho")
        return False
    
def mensagemPadrao():
    global cancelar
    if cancelar:
        atualizar_log("Processamento cancelado!", cor="azul")
        return
    
    mensagem = "Teste Desconsiderando mensagem"


    # mensagem = f"COMUNICADO IMPORTANTE: LIBERAÇÃO DO EMPRÉSTIMO CONSIGNADO (E-CONSIGNADO) PARA FUNCIONÁRIOS\n"
    # mensagem += " \n"
    # mensagem += f"Prezados clientes,\n"
    # mensagem += " \n"
    # mensagem += f"O governo federal liberou a modalidade de empréstimo consignado (e-consignado) para os funcionários, com desconto diretamente na folha de pagamento. Essa medida oferece aos trabalhadores acesso a crédito com taxas de juros reduzidas, facilitando o pagamento por meio do desconto automático em seus salários.\n"
    # mensagem += " \n"
    # mensagem += f"O que é o e-consignado?\n"
    # mensagem += " \n"
    # mensagem += f"O e-consignado é uma linha de crédito pessoal em que as parcelas são descontadas diretamente da folha de pagamento do trabalhador. Essa modalidade oferece juros mais baixos em relação a outras opções de crédito, pois o pagamento é garantido pelo vínculo empregatício.\n"
    # mensagem += " \n"
    # mensagem += f"O que as empresas precisam saber?\n"
    # mensagem += " \n"
    # mensagem += f"• Contratação pelo funcionário: O processo de solicitação e contratação do empréstimo é realizado diretamente pelo trabalhador, por meio da sua CTPS digital, sem necessidade de intermediários.\n"
    # mensagem += f"• Desconto na folha de pagamento: A empresa deve realizar o desconto conforme os valores acordados no contrato do empréstimo.\n"
    # mensagem += f"• Limite de comprometimento da renda: O valor das parcelas não pode ultrapassar 35% da remuneração do trabalhador, conforme a legislação vigente.\n"
    # mensagem += f"• Informação das contratações: As empresas receberão as informações sobre os empréstimos contratados através do DET (Domicílio Eletrônico Trabalhista).\n"
    # mensagem += " \n"
    # mensagem += f"Prazos e regulamentação:\n"
    # mensagem += f"O e-consignado já está disponível e pode ser contratado pelos trabalhadores conforme as regras estabelecidas pelo governo.\n"
    # mensagem += " \n"
    # mensagem += f"Reforçamos a importância de que a empresa garanta que o desconto em folha seja realizado corretamente, evitando inconsistências que possam gerar problemas futuros.\n"
    # mensagem += " \n"
    # mensagem += f"Honorários para Gestão do E-Consignado:\n"
    # mensagem += f"Devido a toda tramitação e responsabilidade envolvidas na operacionalização desse novo serviço, informamos que serão cobrados honorários à parte por colaborador que contratar o e-consignado, conforme orientação que aguardamos do nosso conselho e da associação dos contadores. Os valores correspondentes a esses honorários serão informados oportunamente.\n"
    # mensagem += " \n"
    # mensagem += f"Para mais informações, estamos à disposição para orientá-los!\n"
    # mensagem += " \n"
    # mensagem += f"Atenciosamente,\n"
    # mensagem += f"Equipe Canella e Santos"
    
    return mensagem


def ler_dados_excel(caminho_excel, linha_inicial=2):
    try:
        # Carrega o workbook
        wb = openpyxl.load_workbook(caminho_excel)
        
        # Seleciona a primeira planilha
        sheet = wb.active
        
        # Lista para armazenar os dados
        dados = {}
         
        # Itera pelas linhas da planilha a partir da linha inicial
        for row in sheet.iter_rows(min_row=linha_inicial, values_only=True):
            if row and len(row) >= 4:  # Verifica se a linha tem pelo menos 4 colunas
                codigo, pessoas, nome_contato, nome_grupo = row[:4]
                
                dados[codigo] = {
                    'codigo': codigo,
                    'empresa': pessoas,
                    'nome_contato': nome_contato,
                    'nome_grupo': nome_grupo
                }
            else:
                atualizar_log(f"Linha ignorada por não ter todas as colunas necessárias (esperado 4, encontrado {len(row) if row else 0}): {row}")
        
        if not dados:
            atualizar_log("Nenhum dado válido encontrado no arquivo Excel.", cor="vermelho")
            return None
        
        return dados
    
    except Exception as e:
        atualizar_log(f"Erro ao ler o arquivo Excel: {str(e)}", cor="vermelho")
        return None

def extrair_cod_nome_contatos_e_grupos(dados):
    codigo = []
    nome_contato = []
    nome_grupo = []
    empresas = []
    
    # Iterar sobre o dicionário, onde a chave é o código da empresa
    for cod, info in dados.items():
        codigo.append(cod)  # A chave é o código da empresa
        nome_contato.append(info['nome_contato'])  # Extrair o nome do contato
        nome_grupo.append(info['nome_grupo'])  # Extrair o nome do grupo
        empresas.append(info['empresa'])  # Extrair o nome da empresa
    
    return codigo, empresas, nome_contato, nome_grupo

def inicializar_arquivo_log():
    global log_file_path
    log_dir = os.path.join(os.path.expanduser('~'), 'Documents', 'AutoMessenger_Logs')
    os.makedirs(log_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file_path = os.path.join(log_dir, f"automessenger_log_{timestamp}.txt")
    
    with open(log_file_path, 'w', encoding='utf-8') as f:
        f.write(f"=== Log de Execução - AutoMessenger - {timestamp} ===\n\n")
    
    return log_file_path
        
# Função para selecionar o arquivo Excel
def selecionar_excel():
    arquivo = filedialog.askopenfilename(
        title="Selecione o arquivo Excel",
        filetypes=(("Arquivos Excel", "*.xlsx *.xls"), ("Todos os arquivos", "*.*"))
    )
    if arquivo:
        caminho_excel.set(arquivo)
        atualizar_log(f"Arquivo Excel selecionado: {arquivo}")

# Função para iniciar o processamento dos dados
def iniciar_processamento():
    global cancelar
    cancelar = False
    
    excel = caminho_excel.get()
    if excel:
        try:
            linha_inicial = int(entrada_linha_inicial.get())
            if linha_inicial < 2:
                messagebox.showwarning("Atenção", "A linha inicial deve ser pelo menos 2 (para pular o cabeçalho).")
                return
        except ValueError:
            messagebox.showwarning("Atenção", "A linha inicial deve ser um número inteiro.")
            return
            
        atualizar_log("Iniciando processamento...", cor="azul")
        botao_iniciar.configure(state="disabled")  # Desabilitar o botão para evitar múltiplos cliques
        
        # Inicializa o arquivo de log
        inicializar_arquivo_log()
        atualizar_log(f"Arquivo de log criado em: {log_file_path}")
        
        thread = threading.Thread(target=processar_dados, args=(excel, linha_inicial))
        thread.start()
    else:
        messagebox.showwarning("Atenção", "Por favor, selecione o arquivo Excel.")

# Função do seu código que já está pronta para processar os dados
def processar_dados(excel, linha_inicial):
    global cancelar, progresso

    url_desejada = "https://app.gestta.com.br/attendance/#/chat/contact-list"
    driver = abrir_chrome_com_url(url_desejada)
    
    # Arquivo Excel
    caminho_excel = excel

# Obter o número total de linhas no Excel
    try:
        wb = openpyxl.load_workbook(caminho_excel)
        sheet = wb.active
        total_linhas_excel = sheet.max_row - linha_inicial + 1  # Subtrair 1 para o cabeçalho
        atualizar_log(f"Total de linhas no Excel: {total_linhas_excel}")
    except Exception as e:
        atualizar_log(f"Erro ao obter o total de linhas do Excel: {str(e)}", cor="vermelho")
        total_linhas_excel = 0

    dados = ler_dados_excel(caminho_excel, linha_inicial)
    if not dados:
        atualizar_log("Nenhum dado para processar.", cor="vermelho")
        return
    codigos, empresas, nome_contatos, nome_grupos = extrair_cod_nome_contatos_e_grupos(dados)
    
    # Configuração da barra de progresso
    total_contatos = len(codigos)
    linha_atual_rel = linha_inicial - 2  # Considerando que linha 2 é a primeira linha de dados

    atualizar_log(f"Processando {total_contatos} contatos a partir da linha {linha_inicial} de {total_linhas_excel + 1} totais")
    
    if not driver:
        atualizar_log("Falha ao abrir o Chrome.", cor="vermelho")
        return

    try:
        atualizar_log("Chrome aberto com sucesso.", cor="azul")
        atualizar_log("Aguardando 10 segundos para garantir carregamento completo...")
        time.sleep(10)
        

        #RODAR NOME DE CONTATOS E GRUPOS
        for i, (codigo, empresa, nome_contato, nome_grupo) in enumerate(zip(codigos, empresas, nome_contatos, nome_grupos)):
            if cancelar:
                atualizar_log("Processamento cancelado!", cor="azul")
                return 
            
            # Atualiza a barra de progresso em relação ao total do Excel
            linha_atual = linha_inicial + i
            if total_linhas_excel > 0:
                porcentagem = ((i + 1) / total_contatos) * 100
                
                atualizar_progresso(porcentagem, f"Linha {linha_atual} de {total_linhas_excel + linha_inicial - 1}")
                
            else:
                # Fallback se não conseguiu obter o total de linhas
                porcentagem = ((i + 1) / total_contatos) * 100
                atualizar_progresso(porcentagem, f"{i + 1}/{total_contatos}")
            
            atualizar_log(f"\nProcessando linha {linha_atual}/{total_linhas_excel + linha_inicial - 1}: {codigo} - {empresa}: Contato: {nome_contato}, Grupo: {nome_grupo}\n", cor="azul")

            # Gravar posição atual no log para recuperação
            with open(log_file_path, 'a', encoding='utf-8') as f:
                f.write(f"\n[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] Processando código: {codigo}, empresa: {empresa}, linha Excel: {linha_atual}\n")
                
            mensagem = mensagemPadrao()
                    
            if nome_contato.upper() != "NONE":
                if cancelar:
                    atualizar_log("Processamento cancelado!", cor="azul")
                    return 
                # Encontra a barra de Pesquisa e processa o contato
                if encontrar_e_clicar_barra_contatos(driver, nome_contato, nome_grupo):
                    
                    atualizar_log(f"Contato {nome_contato} na Aba Contato, encontrado e clicado.")
                    try: 
                        # Espera pela area da mensagem
                        area_mensagem = WebDriverWait(driver, 10).until(
                            EC.presence_of_element_located((By.XPATH, '//*[@id="preview-root"]/div[2]'))
                        )
                        
                        if area_mensagem:
                            atualizar_log(f"carregamento Barra de Mensagens 6 segundos ...")
                            # Aguarda a página do chat carregar
                            time.sleep(6)
                                
                            # Foca na barra de mensagem e envia
                            if focar_barra_mensagem_enviar(driver, mensagem):
                                if cancelar:
                                    atualizar_log("Processamento cancelado!", cor="azul")
                                    return

                                atualizar_log(f"\nAviso enviado para {nome_contato}, {codigo} - {empresa}.\n", cor="verde")
                                # Registra o sucesso no log
                                with open(log_file_path, 'a', encoding='utf-8') as f:
                                    f.write(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ✓ Mensagem enviada com sucesso para {nome_contato}\n")
                            else:
                                atualizar_log(f"\nFalha ao enviar mensagem para {nome_contato}\n", cor="vermelho")                       
                                # Registra a falha no log
                                with open(log_file_path, 'a', encoding='utf-8') as f:
                                    f.write(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ✗ FALHA ao enviar mensagem para {nome_contato}\n")
                                
                            time.sleep(5)  # Espera um pouco antes de processar o próximo contato
                                
                            # Clica no botão para voltar à lista de contatos
                            if not focar_pagina_geral(driver):
                                if cancelar:
                                    atualizar_log("Processamento cancelado!", cor="azul")
                                    return
                                
                                atualizar_log("Falha ao voltar para a lista de contatos. Tentando continuar...", cor="vermelho")
                                
                                time.sleep(5)  # Espera adicional após voltar à lista de contatos
                                
                            else:
                                atualizar_log(f"Iniciando novo contato...", cor="azul")
                    
                    except TimeoutException:
                                if cancelar:
                                    atualizar_log("Processamento cancelado!", cor="azul")
                                    return 
                                atualizar_log("\nBug ao tentar clicar no contato na Aba Contato!\n", cor="vermelho")
                                atualizar_log("Clicando na Aba Grupo para solucionar o problema", cor="azul")
                                # focar na aba grupos dos contatos
                                elemento = WebDriverWait(driver, 10).until(
                                    EC.element_to_be_clickable((By.XPATH, '//*[@id="react-tabs-2"]'))
                                )
                                
                                elemento.click()  
                                atualizar_log("Clicando no contato em Grupo.")
                                processar_resultados_busca(driver)     
                                
                                atualizar_log(f"carregamento Barra de Mensagens 6 segundos ...")
                                # Aguarda a página do chat carregar
                                time.sleep(6)
                                    
                                # Foca na barra de mensagem e envia
                                if focar_barra_mensagem_enviar(driver, mensagem):
                                    if cancelar:
                                        atualizar_log("Processamento cancelado!", cor="azul")
                                        return
                                    
    
                                    atualizar_log(f"\nAviso enviado para {nome_contato}, {codigo} - {empresa}.\n", cor="verde")
                                    # Registra o sucesso no log
                                    with open(log_file_path, 'a', encoding='utf-8') as f:
                                        f.write(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ✓ Mensagem enviada com sucesso para {nome_contato} (via grupo)\n")
                                    
                                else:
                                    atualizar_log(f"Falha ao enviar mensagem para {nome_contato}", cor="vermelho")                       
                                    # Registra a falha no log
                                    with open(log_file_path, 'a', encoding='utf-8') as f:
                                        f.write(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ✗ FALHA ao enviar mensagem para {nome_contato}\n")
                                    
                                time.sleep(5)  # Espera um pouco antes de processar o próximo contato
                                    
                                # Clica no botão para voltar à lista de contatos
                                if not focar_pagina_geral(driver):
                                    atualizar_log("Falha ao voltar para a lista de contatos. Tentando continuar...", cor="vermelho")
                                    
                                    time.sleep(5)  # Espera adicional após voltar à lista de contatos
                                    
                                else:
                                    atualizar_log(f"Iniciando novo contato...", cor="azul")
            #GRUPO   
            elif nome_grupo.upper() != "NONE":
                if cancelar:
                    atualizar_log("Processamento cancelado!", cor="azul")
                    return 
                # Encontra a barra de Pesquisa e processa o contato
                if encontrar_e_clicar_barra_contatos(driver, nome_contato, nome_grupo):
                    
                    atualizar_log(f"Contato {nome_grupo} na Aba Grupo, encontrado e clicado.")
                    
                    try:
                        
                        # Espera pela TITULO area da mensagem
                        area_mensagem = WebDriverWait(driver, 10).until(
                            EC.presence_of_element_located((By.XPATH, '//*[@id="ChatHeader"]/div[1]'))
                        )
                        
                        if area_mensagem:
                            
                            atualizar_log(f"carregamento Barra de Mensagens 6 segundos ...")
                            # Aguarda a página do chat carregar
                            time.sleep(6)

                            # Foca na barra de mensagem e envia
                            if focar_barra_mensagem_enviar(driver, mensagem):
                                atualizar_log(f"\nAviso enviado para {nome_grupo}, {codigo} - {empresa}.\n", cor="verde")
                                # Registra o sucesso no log
                                with open(log_file_path, 'a', encoding='utf-8') as f:
                                    f.write(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ✓ Mensagem enviada com sucesso para o grupo {nome_grupo}\n")
                            else:
                                atualizar_log(f"Falha ao enviar mensagem para {nome_grupo}", cor="vermelho")
                                # Registra a falha no log
                                with open(log_file_path, 'a', encoding='utf-8') as f:
                                    f.write(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ✗ FALHA ao enviar mensagem para o grupo {nome_grupo}\n")
                                
                            time.sleep(5)  # Espera um pouco antes de processar o próximo contato
                                
                            # Clica no botão para voltar à lista de contatos
                            if not focar_pagina_geral(driver):
                                atualizar_log("Falha ao voltar para a lista de contatos. Tentando continuar...", cor="vermelho")
                                
                                time.sleep(5)  # Espera adicional após voltar à lista de contatos
                                
                            else:
                                atualizar_log(f"Iniciando novo contato...", cor="azul")
                    except TimeoutException: 
                        if cancelar:
                            atualizar_log("Processamento cancelado!", cor="azul")
                            return 
                        atualizar_log("Bug ao tentar clicar no contato na Aba Grupo!", cor="vermelho")
                        atualizar_log("Clicando na Aba Contato para solucionar o problema", cor="azul")
                        # Clicar aba Contato
                        elemento = WebDriverWait(driver, 10).until(
                            EC.element_to_be_clickable((By.XPATH, '//*[@id="react-tabs-0"]'))
                        )
                        elemento.click()
                        atualizar_log("Clicando no contato em Grupo.")
                        processar_resultados_busca(driver)     
                        
                        atualizar_log(f"carregamento Barra de Mensagens 6 segundos ...")
                        # Aguarda a página do chat carregar
                        time.sleep(6)
                        
                        # Foca na barra de mensagem e envia
                        if focar_barra_mensagem_enviar(driver, mensagem):
                            atualizar_log(f"\nAviso enviado para {nome_grupo}, {codigo} - {empresa}.\n", cor="verde")
                            # Registra o sucesso no log
                            with open(log_file_path, 'a', encoding='utf-8') as f:
                                f.write(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ✓ Mensagem enviada com sucesso para {nome_grupo} (via contato)\n")
                        else:
                            atualizar_log(f"Falha ao enviar mensagem para {nome_grupo}", cor="vermelho")
                            # Registra a falha no log
                            with open(log_file_path, 'a', encoding='utf-8') as f:
                                f.write(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ✗ FALHA ao enviar mensagem para {nome_grupo}\n")
                            
                        time.sleep(5)  # Espera um pouco antes de processar o próximo contato
                            
                        # Clica no botão para voltar à lista de contatos
                        if not focar_pagina_geral(driver):
                            atualizar_log("Falha ao voltar para a lista de contatos. Tentando continuar...", cor="vermelho")
                            
                            time.sleep(5)  # Espera adicional após voltar à lista de contatos
                            
                        else:
                            atualizar_log(f"Iniciando novo contato...", cor="azul")
            else:
                atualizar_log("\nContato Inexistente!", cor="vermelho")
                atualizar_log(f"Pulando pessoa {codigo} - {empresa}", cor="azul")
                atualizar_log("Pulando em 3 Seg...")
                # Registra o pulo no log
                with open(log_file_path, 'a', encoding='utf-8') as f:
                    f.write(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ⚠ Contato inexistente para {codigo} - {empresa}, pulando...\n")
                time.sleep(3)   
                
        # Atualiza a barra de progresso para 100% ao finalizar
        atualizar_progresso(100, f"Concluído - {total_linhas_excel + linha_inicial - 1}/{total_linhas_excel + linha_inicial - 1}")
        with open(log_file_path, 'a', encoding='utf-8') as f:
            f.write(f"\n[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ✓ PROCESSAMENTO FINALIZADO COM SUCESSO!\n")
            
        atualizar_log("Processamento finalizado com sucesso!")
        finalizar_programa()
    except Exception as e:
        atualizar_log(f"Ocorreu um erro inesperado: {str(e)}", cor="vermelho")
        # Registra o erro no log
        with open(log_file_path, 'a', encoding='utf-8') as f:
            f.write(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ✗ ERRO: {str(e)}\n")
            
        botao_iniciar.configure(state="normal")
    
# Função para atualizar a barra de progresso
def atualizar_progresso(valor, texto=""):
    # Normalizar o valor para o intervalo [0, 1] que a barra de progresso espera
    progresso.set(valor / 100)
    progresso_texto.configure(text=texto)
    janela.update_idletasks()  # Atualiza a interface
    
# Função para cancelar o processamento
def cancelar_processamento():
    global cancelar
    cancelar = True
    atualizar_log("Cancelando processamento...", cor="azul")
    with open(log_file_path, 'a', encoding='utf-8') as f:
        f.write(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] ⚠ Processamento CANCELADO pelo usuário\n")
    botao_fechar.configure(state="normal")  # Habilitar o botão de fechar o programa

# Função para cancelar e fechar o programa
def fechar_programa():
    janela.quit()

# Função para finalizar o programa com uma mensagem
def finalizar_programa():
    messagebox.showinfo("Processo Finalizado", "O processamento foi concluído com sucesso!")
    botao_fechar.configure(state="normal")  # Habilitar o botão de fechar o programa
    botao_iniciar.configure(state="normal")  # Reabilitar o botão de iniciar

# Função para abrir o arquivo de log
def abrir_log():
    global log_file_path
    if log_file_path and os.path.exists(log_file_path):
        os.startfile(log_file_path)
    else:
        messagebox.showinfo("Log não disponível", "Não há arquivo de log disponível para esta sessão.")

# Função para atualizar o log na área de texto
def atualizar_log(mensagem, cor=None):
    log_text.configure(state="normal")  # Habilitar edição temporária
    
    # Obter a hora atual
    timestamp = datetime.now().strftime("[%H:%M:%S] ")
    
    # Inserir timestamp e mensagem
    if cor == "vermelho":
        log_text.insert("end", timestamp, "timestamp")
        log_text.insert("end", mensagem + "\n", "vermelho")
    elif cor == "verde":
        log_text.insert("end", timestamp, "timestamp")
        log_text.insert("end", mensagem + "\n", "verde")
    elif cor == "azul":
        log_text.insert("end", timestamp, "timestamp")
        log_text.insert("end", mensagem + "\n", "azul")
    else:
        log_text.insert("end", timestamp, "timestamp")
        log_text.insert("end", mensagem + "\n")
    
    log_text.configure(state="disabled")  # Desabilitar edição novamente
    log_text.see("end")  # Scroll automático para a última linha
    
    # Se o arquivo de log já existir, adiciona a mensagem também
    if log_file_path and os.path.exists(log_file_path):
        with open(log_file_path, 'a', encoding='utf-8') as f:
            f.write(f"{timestamp}{mensagem}\n")

# Função main para encapsular a lógica do programa
def main():
    global janela, caminho_excel, botao_fechar, botao_iniciar, log_text, progresso, progresso_texto, entrada_linha_inicial

    # Criar a janela principal com customtkinter
    janela = ctk.CTk()
    janela.title("Aviso All Clientes AutoMessenger")
    janela.geometry("700x600")
    janela.resizable(True, True)

    # Variáveis para armazenar os caminhos
    caminho_excel = ctk.StringVar()
    progresso = ctk.DoubleVar()

    # Frame para o título
    frame_titulo = ctk.CTkFrame(janela)
    frame_titulo.pack(fill="x", padx=10, pady=10)

    titulo = ctk.CTkLabel(frame_titulo, text="AutoMessenger - Envio Automatizado", font=("Roboto", 16, "bold"))
    titulo.pack(pady=10)

    # Frame para seleção de arquivo
    frame_selecao = ctk.CTkFrame(janela)
    frame_selecao.pack(fill="x", padx=10, pady=5)

    # Label e Botão para selecionar o arquivo Excel
    label_excel = ctk.CTkLabel(frame_selecao, text="Arquivo Excel:")
    label_excel.grid(row=0, column=0, pady=5, padx=5, sticky="w")

    entrada_excel = ctk.CTkEntry(frame_selecao, textvariable=caminho_excel, width=400)
    entrada_excel.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

    botao_excel = ctk.CTkButton(frame_selecao, text="Selecionar Excel", command=selecionar_excel)
    botao_excel.grid(row=0, column=2, padx=5, pady=5)
    
    # Linha inicial para processamento
    label_linha_inicial = ctk.CTkLabel(frame_selecao, text="Iniciar da linha:")
    label_linha_inicial.grid(row=1, column=0, pady=5, padx=5, sticky="w")

    entrada_linha_inicial = ctk.CTkEntry(frame_selecao, width=100)
    entrada_linha_inicial.grid(row=1, column=1, padx=5, pady=5, sticky="w")
    entrada_linha_inicial.insert(0, "2")  # Valor padrão
    
    frame_selecao.grid_columnconfigure(1, weight=1)  # Fazer a coluna do meio expandir

    # Frame para botões de ação
    frame_botoes = ctk.CTkFrame(janela)
    frame_botoes.pack(fill="x", padx=10, pady=5)

    # Botões para operações
    botao_iniciar = ctk.CTkButton(frame_botoes, text="Iniciar Processamento", command=iniciar_processamento, fg_color="#28a745", hover_color="#218838")
    botao_iniciar.pack(side="left", padx=5, pady=10, expand=True, fill="x")

    botao_cancelar = ctk.CTkButton(frame_botoes, text="Cancelar Processamento", command=cancelar_processamento, fg_color="#dc3545", hover_color="#c82333")
    botao_cancelar.pack(side="left", padx=5, pady=10, expand=True, fill="x")

    botao_fechar = ctk.CTkButton(frame_botoes, text="Fechar Programa", command=fechar_programa, state="disabled", fg_color="#6c757d", hover_color="#5a6268")
    botao_fechar.pack(side="left", padx=5, pady=10, expand=True, fill="x")
    
    botao_abrir_log = ctk.CTkButton(frame_botoes, text="Abrir Log", command=abrir_log, fg_color="#17a2b8", hover_color="#138496")
    botao_abrir_log.pack(side="left", padx=5, pady=10, expand=True, fill="x")

    # Frame para a barra de progresso
    frame_progresso = ctk.CTkFrame(janela)
    frame_progresso.pack(fill="x", padx=10, pady=5)
    
    label_progresso = ctk.CTkLabel(frame_progresso, text="Progresso:")
    label_progresso.pack(side="left", padx=5)
    
    barra_progresso = ctk.CTkProgressBar(frame_progresso, variable=progresso, width=500)
    barra_progresso.pack(side="left", padx=5, fill="x", expand=True)
    barra_progresso.set(0)
    
    progresso_texto = ctk.CTkLabel(frame_progresso, text="0/0")
    progresso_texto.pack(side="left", padx=5)

    # Frame para o log
    frame_log = ctk.CTkFrame(janela)
    frame_log.pack(pady=10, padx=10, fill="both", expand=True)

    # Label para o log
    label_log = ctk.CTkLabel(frame_log, text="Log de execução:")
    label_log.pack(anchor="w", padx=5, pady=5)

    # Área de texto para o log com barra de rolagem
    log_text = ctk.CTkTextbox(frame_log, wrap="word", height=250, width=650)
    log_text.pack(fill="both", expand=True, padx=5, pady=5)
    
    # Configurar tags para cores
    log_text.tag_config("vermelho", foreground="red")
    log_text.tag_config("verde", foreground="green")
    log_text.tag_config("azul", foreground="blue")
    log_text.tag_config("timestamp", foreground="gray")

    # Inicializar o log com uma mensagem
    atualizar_log("Bem-vindo ao AutoMessenger! Selecione um arquivo Excel e clique em 'Iniciar Processamento'.", cor="azul")
    
    # Rodapé
    frame_rodape = ctk.CTkFrame(janela, fg_color="transparent")
    frame_rodape.pack(fill="x", padx=10, pady=5)
    
    label_versao = ctk.CTkLabel(frame_rodape, text="v1.0 | Desenvolvido por Hugo L. Almeida - Equipe de TI", text_color="gray")
    label_versao.pack(side="right", padx=5, pady=5)
    
    # Iniciar o loop da interface
    janela.mainloop()

# Garantir que o código só execute se este arquivo for o principal
if __name__ == '__main__':
    main()